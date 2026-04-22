"""
=============================================================================
BATTERY SOC LIVE PREDICTION SYSTEM  ·  Multi-Model Edition
=============================================================================
Features:
  1. AUTO MODE   - Reads live sensor data from Arduino via Serial port
  2. MANUAL MODE - Enter SOC value manually to test ESP32 relay
  3. Battery type selector — each battery loads its own trained model:
       Li-Ion / LFP  →  3V_battery_soc_model.pkl
       SLA 12 V      →  13_battery_soc_model.pkl
       SLA 24 V      →  24_battery_soc_model.pkl
       NiMH          →  battery_soc_model.pkl  (generic fallback)
  4. Saves all data to live Excel file (served via http.server 8000)
  5. ESP32 local web endpoint for SOC polling (both manual & auto)
  6. Self-debugging error reporter with detailed logs
  7. Background HTTP server (no extra terminal needed)
=============================================================================
REQUIRED INSTALLS:
  pip install pyserial joblib scikit-learn openpyxl flask pandas numpy
=============================================================================
ARDUINO SETUP:
  Serial.print(voltage); Serial.print(",");
  Serial.print(current); Serial.print(",");
  Serial.print(temperature); Serial.print(",");
  Serial.println(time_s);
=============================================================================
ESP32 ENDPOINT:
  http://YOUR_PC_IP:8000/soc_feed.json
  Returns: {"soc":75.34,"mode":"auto","battery":"SLA 12V","relay_on":false,...}
=============================================================================
"""

im sys
import os
import json
import time
import serial
import serial.tools.list_ports
import threading
import logging
import traceback
import http.server
import socketserver
import functools
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
import joblib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────── PATHS ────────────────────────────────────────
BASE_DIR     = Path(__file__).parent

MODEL_DIR    = Path(r"D:\SOC_ML_Project\modal")
EXCEL_FILE   = Path(r"D:\SOC_ML_Project\data\prediction\predicted_output.xlsx")
LOG_FILE     = Path(r"D:\SOC_ML_Project\logs\debug.log")
SOC_JSON     = Path(r"D:\SOC_ML_Project\soc_feed.json")

# ─────────────────────────── LOGGING ──────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger("BatterySOC")

# ─────────────────────── BATTERY PROFILES ─────────────────────────────────
BATTERY_PROFILES = {
    "1": {
        "name":         "Li-Ion (18650 / Laptop)",
        "model_prefix": "3V",
        "min_v": 3.0,  "max_v": 4.2,
        "min_soc": 0,  "max_soc": 100,
        "temp_warn": 45, "temp_crit": 60,
        "capacity_ah": 2.5,
    },
    "2": {
        "name":         "LFP (LiFePO4)",
        "model_prefix": "3V",
        "min_v": 2.5,  "max_v": 3.65,
        "min_soc": 0,  "max_soc": 100,
        "temp_warn": 55, "temp_crit": 75,
        "capacity_ah": 10.0,
    },
    "3": {
        "name":         "NiMH",
        "model_prefix": None,
        "min_v": 1.0,  "max_v": 1.45,
        "min_soc": 10, "max_soc": 100,
        "temp_warn": 45, "temp_crit": 55,
        "capacity_ah": 2.0,
    },
    "4": {
        "name":         "SLA 12V (Sealed Lead Acid)",
        "model_prefix": "13",
        "min_v": 10.5, "max_v": 13.8,
        "min_soc": 20, "max_soc": 100,
        "temp_warn": 40, "temp_crit": 50,
        "capacity_ah": 7.0,
    },
    "5": {
        "name":         "SLA 24V (Sealed Lead Acid)",
        "model_prefix": "24",
        "min_v": 21.0, "max_v": 27.6,
        "min_soc": 20, "max_soc": 100,
        "temp_warn": 40, "temp_crit": 50,
        "capacity_ah": 7.0,
    },
}

# ─────────────────────────── GLOBALS ──────────────────────────────────────
_current_soc  = 0.0
_current_mode = "idle"
_running      = False
_soc_lock     = threading.Lock()
_http_port    = 8000


# ═══════════════════════════════════════════════════════════════════════════
#  SELF-DEBUGGING HELPER
# ═══════════════════════════════════════════════════════════════════════════

class SelfDebugger:

    @staticmethod
    def capture(exc, context=""):
        tb    = traceback.format_exc()
        etype = type(exc).__name__
        msg   = str(exc)
        diag  = SelfDebugger._diagnose(etype, msg)
        full  = (
            f"\n{'━'*60}\n"
            f"  ⚠  ERROR in {context}\n"
            f"{'━'*60}\n"
            f"  Type    : {etype}\n"
            f"  Message : {msg}\n"
            f"  Fix Tip : {diag}\n"
            f"{'━'*60}\n"
            f"  Full traceback saved to: {LOG_FILE}\n"
            f"{'━'*60}"
        )
        log.error(full)
        log.debug(tb)
        return full

    @staticmethod
    def _diagnose(etype, msg):
        rules = [
            ("FileNotFoundError", "battery_soc_model.pkl",
             "Model .pkl not found in D:\\SOC_ML_Project\\modal\\ — check prefix or re-run training."),
            ("FileNotFoundError", "",
             "Check that all file paths are correct and files exist."),
            ("SerialException", "could not open port",
             "Port busy. Close Arduino IDE Serial Monitor and retry."),
            ("SerialException", "Access is denied",
             "Arduino IDE Serial Monitor is open — close it first."),
            ("SerialException", "No serial ports",
             "Arduino not detected. Check USB cable and CH340/FTDI driver."),
            ("SerialException", "",
             "Arduino not connected or driver missing. Install CH340 driver."),
            ("UnicodeDecodeError", "",
             "Serial data has non-UTF8 bytes. Check Arduino baud rate."),
            ("ValueError", "could not convert",
             "Arduino sent non-numeric data. Verify Serial.print() format."),
            ("ModuleNotFoundError", "serial",   "Run: pip install pyserial"),
            ("ModuleNotFoundError", "openpyxl", "Run: pip install openpyxl"),
            ("ModuleNotFoundError", "joblib",   "Run: pip install joblib"),
            ("OSError", "10048",
             "Port 8000 already in use. Kill the process or change _http_port."),
            ("OSError", "Address already in use",
             "Port 8000 already in use. Change _http_port at top of script."),
            ("PermissionError", "",
             "No write permission. Run as Administrator or change BASE_DIR."),
            ("KeyboardInterrupt", "", "User stopped the program."),
        ]
        for ep, mp, tip in rules:
            if ep in etype and mp.lower() in msg.lower():
                return tip
        return "Unexpected error — see full traceback in debug.log."


def safe_run(fn, *args, context="", **kwargs):
    try:
        return fn(*args, **kwargs)
    except Exception as exc:
        SelfDebugger.capture(exc, context or fn.__name__)
        return None


# ═══════════════════════════════════════════════════════════════════════════
#  MULTI-MODEL LOADER
# ═══════════════════════════════════════════════════════════════════════════

def _model_paths(prefix):
    if prefix:
        return (
            MODEL_DIR / f"{prefix}_battery_soc_model.pkl",
            MODEL_DIR / f"{prefix}_feature_columns.pkl",
        )
    return (
        MODEL_DIR / "battery_soc_model.pkl",
        MODEL_DIR / "feature_columns.pkl",
    )


def _model_label(profile):
    mp, _ = _model_paths(profile.get("model_prefix"))
    return mp.name


def load_model_for_battery(profile):
    prefix = profile.get("model_prefix")
    mp, fp = _model_paths(prefix)

    if not mp.exists():
        log.warning(f"Model not found: {mp}  →  falling back to generic model.")
        print(f"\n  ⚠  {mp.name} not found — falling back to battery_soc_model.pkl\n")
        mp, fp = _model_paths(None)

    if not mp.exists():
        raise FileNotFoundError(
            f"No model found at {mp}\n"
            "  ➜  Run the training script first to generate model files."
        )
    if not fp.exists():
        raise FileNotFoundError(f"Feature-columns file not found: {fp}")

    model     = joblib.load(mp)
    feat_cols = joblib.load(fp)
    log.info(f"Loaded model : {mp.name}")
    log.info(f"Features     : {feat_cols}")
    return model, feat_cols


def preload_all_models():
    cache = {}
    print("\n  📦  Pre-loading battery models …")
    for key, profile in BATTERY_PROFILES.items():
        prefix = profile.get("model_prefix")
        mp, fp = _model_paths(prefix)
        target_mp, target_fp = (mp, fp) if (mp.exists() and fp.exists()) else _model_paths(None)
        label = target_mp.name

        if target_mp.exists() and target_fp.exists():
            try:
                cache[key] = (joblib.load(target_mp), joblib.load(target_fp))
                tag = "✅" if mp.exists() else "⚡ fallback"
                print(f"    {tag}  [{key}] {profile['name']:<32} ← {label}")
                log.info(f"Pre-loaded [{key}] {profile['name']} using {label}")
            except Exception as exc:
                print(f"    ❌  [{key}] {profile['name']:<32} ← LOAD ERROR: {exc}")
                log.error(f"Failed to load {target_mp}: {exc}")
        else:
            print(f"    ⚠   [{key}] {profile['name']:<32} ← NO MODEL FOUND")
            log.warning(f"No model available for profile {key}.")
    print()
    return cache


# ═══════════════════════════════════════════════════════════════════════════
#  PREDICTION
# ═══════════════════════════════════════════════════════════════════════════

def predict_soc(model, feat_cols, voltage, current, temp, time_s):
    row = pd.DataFrame([[voltage, current, temp, time_s]], columns=feat_cols)
    soc = float(model.predict(row)[0])
    return round(max(0.0, min(100.0, soc)), 2)


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════

_excel_lock = threading.Lock()
_excel_cols = [
    "Timestamp", "Mode", "Battery Type", "Model Used",
    "Voltage(V)", "Current(A)", "Temp(°C)", "Time(s)",
    "SOC(%)", "Health", "Notes"
]

def _excel_header_style(ws):
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin     = Side(style="thin", color="AAAAAA")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, name in enumerate(_excel_cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = max(16, len(name) + 4)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def _ensure_excel():
    if not EXCEL_FILE.exists():
        EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOC Live Data"
        _excel_header_style(ws)
        wb.save(EXCEL_FILE)
        log.info(f"Created Excel: {EXCEL_FILE}")

def _soc_health_label(soc):
    if soc >= 80: return "✅ Good"
    if soc >= 40: return "⚡ Normal"
    if soc >= 20: return "⚠️ Low"
    return "🔴 Critical"

def append_to_excel(mode, profile, voltage=None, current=None,
                    temp=None, time_s=None, soc=None, notes=""):
    with _excel_lock:
        try:
            _ensure_excel()
            wb       = openpyxl.load_workbook(EXCEL_FILE)
            ws       = wb.active
            next_row = ws.max_row + 1
            thin     = Side(style="thin", color="DDDDDD")
            border   = Border(left=thin, right=thin, top=thin, bottom=thin)
            row_fill = (
                PatternFill("solid", fgColor="EBF5FB")
                if next_row % 2 == 0
                else PatternFill("solid", fgColor="FFFFFF")
            )
            values = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                mode.upper(),
                profile["name"],
                _model_label(profile),
                voltage  if voltage  is not None else "—",
                current  if current  is not None else "—",
                temp     if temp     is not None else "—",
                time_s   if time_s   is not None else "—",
                soc      if soc      is not None else "—",
                _soc_health_label(soc or 0),
                notes,
            ]
            for ci, val in enumerate(values, 1):
                cell           = ws.cell(row=next_row, column=ci, value=val)
                cell.border    = border
                cell.fill      = row_fill
                cell.alignment = Alignment(horizontal="center")
            wb.save(EXCEL_FILE)
        except Exception as exc:
            SelfDebugger.capture(exc, "append_to_excel")


# ═══════════════════════════════════════════════════════════════════════════
#  SOC JSON FEED
# ═══════════════════════════════════════════════════════════════════════════

def update_soc_json(soc, mode, battery_name, model_used="",
                    voltage=None, current=None, temp=None):
    global _current_soc
    with _soc_lock:
        _current_soc = soc
        payload = {
            "soc":        round(soc, 2),
            "mode":       mode,
            "battery":    battery_name,
            "model_used": model_used,
            "voltage":    voltage,
            "current":    current,
            "temp":       temp,
            "ts":         datetime.now().isoformat(),
            "relay_on":   soc < 20.0,
        }
        SOC_JSON.parent.mkdir(parents=True, exist_ok=True)
        SOC_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    log.debug(f"SOC JSON updated: SOC={soc:.2f}% mode={mode} model={model_used}")


# ═══════════════════════════════════════════════════════════════════════════
#  HTTP SERVER
# ═══════════════════════════════════════════════════════════════════════════

class QuietHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        super().end_headers()

    def log_message(self, fmt, *args):
        log.debug(f"HTTP: {fmt % args}")

def start_http_server(port=8000):
    global _http_port
    _http_port = port

    def _serve():
        handler = functools.partial(QuietHandler)
        with socketserver.TCPServer(("", port), handler) as httpd:
            httpd.allow_reuse_address = True
            log.info(f"HTTP server started on port {port}")
            print(f"\n  🌐  Server         →  http://localhost:{port}/")
            print(f"  📊  Dashboard      →  http://localhost:{port}/view_csv.html")
            print(f"  🤖  ESP32 SOC feed →  http://localhost:{port}/soc_feed.json")
            print(f"  📁  Excel file     →  http://localhost:{port}/predicted_output.xlsx\n")
            httpd.serve_forever()

    threading.Thread(target=_serve, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════════
#  ARDUINO SERIAL AUTO-DETECTION
# ═══════════════════════════════════════════════════════════════════════════

_ARDUINO_KEYWORDS = [
    "arduino", "ch340", "ch341", "ftdi", "usb serial", "usb-serial",
    "uart", "cp210", "cp2102", "cp2104", "uno", "mega", "nano",
    "genuino", "wch", "prolific", "usb2serial",
]

def _score_port(pi):
    desc  = (pi.description or "").lower()
    hwid  = (pi.hwid or "").lower()
    score = sum(10 for kw in _ARDUINO_KEYWORDS if kw in desc or kw in hwid)
    for vid in {"2341", "1a86", "0403", "10c4", "067b", "04d8"}:
        if vid in hwid:
            score += 20
    import re
    m = re.search(r"COM(\d+)", pi.device)
    if m and int(m.group(1)) <= 10:
        score += 2
    return score

def auto_detect_port(baud_rates=(9600, 115200, 57600, 19200)):
    all_ports = list(serial.tools.list_ports.comports())
    if not all_ports:
        raise serial.SerialException(
            "No serial ports found. Check USB cable and driver installation."
        )

    sep    = "─" * 55
    scored = sorted(all_ports, key=_score_port, reverse=True)

    print(f"\n{sep}")
    print("  AUTO-DETECTING Arduino port …")
    print(sep)
    for p in scored:
        s = _score_port(p)
        m = "★" if s >= 20 else ("·" if s > 0 else " ")
        print(f"  {m} {p.device:<10} {p.description}")
    print(sep)

    candidates = [p for p in scored if _score_port(p) > 0] or scored
    for pi in candidates:
        for baud in baud_rates:
            log.debug(f"Probing {pi.device} @ {baud}")
            try:
                ser = serial.Serial(pi.device, baud, timeout=2)
                time.sleep(2)
                ser.reset_input_buffer()
                line = ser.readline().decode("utf-8", errors="replace").strip()
                ser.close()
                if line and "," in line:
                    _ = [float(x.strip()) for x in line.split(",")]
                    print(f"\n  ✅  Arduino detected → {pi.device} @ {baud} baud")
                    print(f"  Sample data : {line}")
                    log.info(f"Auto-detected: {pi.device} @ {baud}")
                    return pi.device, baud
            except (serial.SerialException, ValueError, OSError) as e:
                log.debug(f"  {pi.device} @ {baud}: {e}")

    print("\n  ⚠  Could not auto-validate. Select manually:\n")
    for i, p in enumerate(scored, 1):
        tag = " ← likely Arduino" if _score_port(p) >= 20 else ""
        print(f"    [{i}] {p.device:<10} {p.description}{tag}")

    while True:
        choice = input("\n  Select port number (or Enter to cancel): ").strip()
        if choice == "":
            return None, None
        if choice.isdigit() and 1 <= int(choice) <= len(scored):
            p     = scored[int(choice) - 1]
            rates = [9600, 19200, 38400, 57600, 115200]
            print("\n  Baud rates: " + "  ".join(f"[{i+1}] {r}" for i, r in enumerate(rates)))
            bc   = input("  Select baud [Enter = 9600]: ").strip()
            baud = rates[int(bc) - 1] if bc.isdigit() and 1 <= int(bc) <= len(rates) else 9600
            return p.device, baud
        print("  Invalid. Try again.")


# ═══════════════════════════════════════════════════════════════════════════
#  AUTO MODE
# ═══════════════════════════════════════════════════════════════════════════

def auto_mode(model, feat_cols, profile):
    global _running, _current_mode
    _current_mode = "auto"
    ml = _model_label(profile)

    port, baud = safe_run(auto_detect_port, context="auto_detect_port") or (None, None)
    if port is None:
        print("\n  ❌  No Arduino port found or selected. Returning to menu.\n")
        return

    print(f"\n  📡  Connecting to Arduino on {port} @ {baud} baud …")
    print(f"  🧠  Model : {ml}")
    print("  Press Ctrl+C to stop.\n")

    try:
        ser = serial.Serial(port, baud, timeout=2)
        time.sleep(2)
        ser.reset_input_buffer()
        log.info(f"Serial connected: {port} @ {baud}")
    except Exception as exc:
        SelfDebugger.capture(exc, "serial.Serial open")
        return

    start_time   = time.time()
    sample_count = 0
    parse_errors = 0
    MAX_ERRORS   = 10

    _running = True
    try:
        while _running:
            try:
                raw = ser.readline()
                if not raw:
                    continue
                line = raw.decode("utf-8", errors="replace").strip()
                if not line:
                    continue

                parts = line.split(",")
                if len(parts) != 4:
                    log.warning(f"Bad format: {line!r}")
                    parse_errors += 1
                    if parse_errors >= MAX_ERRORS:
                        print(f"  ⚠  {MAX_ERRORS} parse errors. Expected: voltage,current,temp,time_s")
                    continue

                voltage, current, temp, _ = [float(p.strip()) for p in parts]
                elapsed = round(time.time() - start_time, 1)

                notes = ""
                if not (profile["min_v"] <= voltage <= profile["max_v"]):
                    notes = (
                        f"⚠ V={voltage}V out of range "
                        f"[{profile['min_v']}–{profile['max_v']}V]"
                    )
                    log.warning(notes)
                if temp >= profile["temp_warn"]:
                    notes += f" 🌡 High temp {temp}°C"

                soc = predict_soc(model, feat_cols, voltage, current, temp, elapsed)
                sample_count += 1
                parse_errors  = 0

                update_soc_json(soc, "auto", profile["name"], ml, voltage, current, temp)
                append_to_excel("auto", profile, voltage, current, temp, elapsed, soc, notes)

                bar = "█" * int(soc / 5) + "░" * (20 - int(soc / 5))
                print(
                    f"  [{bar}] SOC: {soc:6.2f}%  |  "
                    f"V:{voltage:.3f}V  I:{current:.3f}A  T:{temp:.1f}°C  {notes}"
                )

            except ValueError as ve:
                SelfDebugger.capture(ve, "serial parse")
                parse_errors += 1
            except serial.SerialException as se:
                SelfDebugger.capture(se, "serial readline")
                break
            except KeyboardInterrupt:
                break

    finally:
        _running = False
        ser.close()
        print(f"\n  ✅  Auto mode stopped.  Samples: {sample_count}")
        log.info(f"Auto mode ended. Samples: {sample_count}")


# ═══════════════════════════════════════════════════════════════════════════
#  MANUAL MODE
# ═══════════════════════════════════════════════════════════════════════════

def manual_mode(profile):
    global _current_mode
    _current_mode = "manual"
    ml = _model_label(profile)

    print("\n" + "═" * 60)
    print("  🔧  MANUAL SOC MODE  —  ESP32 Relay Test")
    print(f"  🧠  Model     : {ml}")
    print(f"  🔋  Battery   : {profile['name']}")
    print(f"  📡  ESP32 URL : http://YOUR_PC_IP:{_http_port}/soc_feed.json")
    print("  Type a SOC value (0–100) and press Enter.  'q' to quit.")
    print("═" * 60)

    while True:
        raw = input("\n  Enter SOC (0-100) or 'q': ").strip()
        if raw.lower() == "q":
            break
        try:
            soc = float(raw)
            if not (0 <= soc <= 100):
                print("  ⚠  SOC must be 0–100.")
                continue

            relay = "ON  🔴 (protection active)" if soc < 20 else "OFF ✅"
            update_soc_json(soc, "manual", profile["name"], ml)
            append_to_excel("manual", profile, soc=soc,
                            notes=f"Manual test | Relay: {relay}")

            print(f"\n  ✅  SOC set    : {soc:.2f}%")
            print(f"  🔌  Relay      : {relay}")
            print(f"  🧠  Model      : {ml}")
            print(f"  📡  Feed URL   : http://localhost:{_http_port}/soc_feed.json")

        except ValueError:
            print("  ❌  Invalid. Enter a number like 75 or 23.5")

    _current_mode = "idle"
    print("\n  Exited manual mode.\n")


# ═══════════════════════════════════════════════════════════════════════════
#  BATTERY SELECTOR
# ═══════════════════════════════════════════════════════════════════════════

def select_battery(model_cache):
    print("\n" + "═" * 60)
    print("  🔋  SELECT BATTERY TYPE")
    print("═" * 60)
    for k, v in BATTERY_PROFILES.items():
        mp, _ = _model_paths(v.get("model_prefix"))
        status = "✅ ready" if mp.exists() else "⚠  fallback to generic"
        print(f"  [{k}] {v['name']}")
        print(f"       Voltage : {v['min_v']}–{v['max_v']} V  |  Capacity : {v['capacity_ah']} Ah")
        print(f"       Model   : {mp.name}  [{status}]")
        print(f"       Temp    : warn >{v['temp_warn']}°C  |  crit >{v['temp_crit']}°C")
        print()

    while True:
        choice = input("  Enter battery number: ").strip()
        if choice not in BATTERY_PROFILES:
            print("  Invalid. Try again.")
            continue

        profile = BATTERY_PROFILES[choice]

        if choice in model_cache:
            model, feat_cols = model_cache[choice]
            print(f"\n  ✅  {profile['name']}  (model from cache: {_model_label(profile)})")
        else:
            print(f"\n  ⏳  Loading {_model_label(profile)} …")
            result = safe_run(load_model_for_battery, profile,
                              context="load_model_for_battery")
            if result is None:
                print("  ❌  Model load failed — AUTO mode unavailable for this battery.")
                model, feat_cols = None, None
            else:
                model, feat_cols = result
                model_cache[choice] = (model, feat_cols)
                print(f"  ✅  {profile['name']}  ← {_model_label(profile)}")

        log.info(f"Battery selected: {profile['name']}  model: {_model_label(profile)}")
        return profile, model, feat_cols


# ═══════════════════════════════════════════════════════════════════════════
#  HTML DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════

_HTML_VIEWER = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Battery SOC Live Dashboard</title>
    <style>
        :root{--primary:#1a4a8a;--accent:#00c6ff;--green:#27ae60;--red:#e74c3c;
              --bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--sub:#8b949e;}
        *{box-sizing:border-box;margin:0;padding:0;}
        body{font-family:'Segoe UI',sans-serif;background:var(--bg);color:var(--text);padding:20px;}
        h1{text-align:center;font-size:1.8em;margin-bottom:5px;
           background:linear-gradient(90deg,var(--accent),#0072ff);
           -webkit-background-clip:text;-webkit-text-fill-color:transparent;}
        .subtitle{text-align:center;color:var(--sub);margin-bottom:20px;font-size:.9em;}
        .soc-card{display:flex;gap:16px;flex-wrap:wrap;justify-content:center;margin-bottom:20px;}
        .metric{background:var(--card);border:1px solid var(--border);border-radius:12px;
                padding:16px 24px;text-align:center;min-width:140px;}
        .metric .label{color:var(--sub);font-size:.8em;margin-bottom:6px;}
        .metric .value{font-size:2em;font-weight:700;color:var(--accent);}
        .metric .value.green{color:var(--green);}.metric .value.red{color:var(--red);}
        .metric .value.small{font-size:.9em;}
        .bar-wrap{background:#21262d;border-radius:8px;height:24px;overflow:hidden;
                  margin:16px auto;max-width:600px;}
        .bar{height:100%;border-radius:8px;transition:width .6s ease,background .3s;
             background:linear-gradient(90deg,#0072ff,#00c6ff);}
        .status{text-align:center;font-size:.9em;color:var(--sub);margin-bottom:20px;}
        .tag{display:inline-block;padding:3px 10px;border-radius:20px;font-size:.8em;
             background:#21262d;border:1px solid var(--border);margin:2px;}
        #toolbar{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px;
                 align-items:center;justify-content:space-between;}
        #searchInput{background:var(--card);border:1px solid var(--border);color:var(--text);
                     padding:8px 14px;border-radius:8px;font-size:.9em;flex:1;min-width:180px;}
        #searchInput:focus{outline:none;border-color:var(--accent);}
        #downloadBtn{padding:8px 18px;background:var(--green);color:white;border:none;
                     border-radius:8px;cursor:pointer;font-size:.9em;}
        #refreshBtn{padding:8px 18px;background:var(--primary);color:white;border:none;
                    border-radius:8px;cursor:pointer;font-size:.9em;}
        .table-wrapper{overflow-x:auto;border-radius:10px;border:1px solid var(--border);}
        table{width:100%;border-collapse:collapse;font-size:.88em;}
        thead th{background:#21262d;color:var(--accent);padding:10px 14px;text-align:left;
                 cursor:pointer;white-space:nowrap;border-bottom:1px solid var(--border);}
        thead th:hover{background:#2d333b;}
        thead th.asc::after{content:' ▲';}thead th.desc::after{content:' ▼';}
        tbody tr:hover{background:#1c2128;}
        td{padding:8px 14px;border-bottom:1px solid var(--border);white-space:nowrap;}
        #pagination{display:flex;justify-content:center;gap:10px;margin-top:16px;flex-wrap:wrap;}
        #pagination button{padding:6px 14px;border:1px solid var(--border);border-radius:6px;
                           background:var(--card);color:var(--text);cursor:pointer;}
        #pagination button:hover{border-color:var(--accent);}
        #pagination button:disabled{opacity:.3;cursor:not-allowed;}
        #pageInfo{color:var(--sub);font-size:.9em;line-height:2;}
    </style>
</head>
<body>
    <h1>⚡ Battery SOC Live Dashboard</h1>
    <p class="subtitle">Multi-model ML predictions · auto-refresh every 3 s</p>
    <div class="soc-card">
        <div class="metric"><div class="label">Current SOC</div>
            <div class="value" id="socVal">—</div></div>
        <div class="metric"><div class="label">Mode</div>
            <div class="value" id="modeVal" style="font-size:1.2em">—</div></div>
        <div class="metric"><div class="label">Battery</div>
            <div class="value" id="batVal" style="font-size:1em">—</div></div>
        <div class="metric"><div class="label">Relay</div>
            <div class="value" id="relayVal">—</div></div>
        <div class="metric"><div class="label">Model</div>
            <div class="value small" id="modelVal">—</div></div>
        <div class="metric"><div class="label">Voltage</div>
            <div class="value small" id="voltVal">—</div></div>
        <div class="metric"><div class="label">Temp</div>
            <div class="value small" id="tempVal">—</div></div>
    </div>
    <div class="bar-wrap"><div class="bar" id="socBar" style="width:0%"></div></div>
    <div class="status">
        <span class="tag" id="liveTag">● LIVE</span>
        <span class="tag" id="tsTag">—</span>
        <span class="tag" id="statsTag">—</span>
    </div>
    <div id="toolbar">
        <input type="text" id="searchInput" placeholder="🔍 Search table...">
        <button id="refreshBtn" onclick="loadTable()">↻ Refresh</button>
        <button id="downloadBtn">⬇ Download CSV</button>
    </div>
    <div class="table-wrapper">
        <table><thead id="tableHead"></thead><tbody id="tableBody"></tbody></table>
    </div>
    <div id="pagination">
        <button id="prevBtn" onclick="changePage(-1)">← Prev</button>
        <span id="pageInfo"></span>
        <button id="nextBtn" onclick="changePage(1)">Next →</button>
    </div>
<script>
const JSON_URL='/soc_feed.json';
let allRows=[],filteredRows=[],headers=[],currentPage=1,rowsPerPage=50;
let sortCol=-1,sortAsc=true,csvText='';
async function fetchSOC(){
    try{
        const r=await fetch(JSON_URL+'?_='+Date.now());
        const d=await r.json();
        const soc=d.soc??0;
        document.getElementById('socVal').textContent=soc.toFixed(1)+'%';
        document.getElementById('modeVal').textContent=(d.mode||'—').toUpperCase();
        document.getElementById('batVal').textContent=d.battery||'—';
        document.getElementById('modelVal').textContent=(d.model_used||'—').replace('_battery_soc_model.pkl','');
        document.getElementById('voltVal').textContent=d.voltage!=null?d.voltage.toFixed(3)+' V':'—';
        document.getElementById('tempVal').textContent=d.temp!=null?d.temp.toFixed(1)+' °C':'—';
        const rv=document.getElementById('relayVal');
        rv.textContent=d.relay_on?'ON 🔴':'OFF ✅';
        rv.className='value '+(d.relay_on?'red':'green');
        const bar=document.getElementById('socBar');
        bar.style.width=soc+'%';
        bar.style.background=soc<20?'linear-gradient(90deg,#e74c3c,#ff6b6b)':
            soc<50?'linear-gradient(90deg,#f39c12,#f1c40f)':
            'linear-gradient(90deg,#0072ff,#00c6ff)';
        document.getElementById('tsTag').textContent='🕒 '+new Date(d.ts).toLocaleTimeString();
        document.getElementById('liveTag').textContent='● LIVE';
        document.getElementById('liveTag').style.color='#27ae60';
    }catch(e){
        document.getElementById('liveTag').textContent='● OFFLINE';
        document.getElementById('liveTag').style.color='#e74c3c';
    }
}
setInterval(fetchSOC,3000);fetchSOC();
async function loadTable(){
    try{
        const r=await fetch('/predicted_output.csv?_='+Date.now());
        if(!r.ok)throw new Error('no csv');
        csvText=await r.text();parseCSV(csvText);
    }catch(e){
        document.getElementById('tableBody').innerHTML=
        '<tr><td colspan="11" style="text-align:center;color:#8b949e;padding:30px">Waiting for data…</td></tr>';
    }
}
function parseCSV(text){
    const lines=text.trim().split('\\n').map(l=>l.split(',').map(c=>c.trim().replace(/^"|"$/g,'')));
    if(lines.length<2)return;
    headers=lines[0];allRows=lines.slice(1);filteredRows=[...allRows];currentPage=1;
    renderHeaders();renderTable();
    document.getElementById('statsTag').textContent=allRows.length+' rows';
    document.getElementById('downloadBtn').onclick=()=>{
        const b=new Blob([csvText],{type:'text/csv'});
        const a=document.createElement('a');a.href=URL.createObjectURL(b);
        a.download='predicted_output.csv';a.click();
    };
}
function renderHeaders(){
    const thead=document.getElementById('tableHead');thead.innerHTML='';
    const tr=document.createElement('tr');
    headers.forEach((h,i)=>{const th=document.createElement('th');th.textContent=h;
        th.onclick=()=>sortTable(i,th);tr.appendChild(th);});
    thead.appendChild(tr);
}
function renderTable(){
    const tbody=document.getElementById('tableBody');tbody.innerHTML='';
    const start=(currentPage-1)*rowsPerPage;
    filteredRows.slice(start,start+rowsPerPage).forEach(row=>{
        const tr=document.createElement('tr');
        headers.forEach((_,i)=>{const td=document.createElement('td');
            td.textContent=row[i]??'';tr.appendChild(td);});
        tbody.appendChild(tr);
    });
    const total=Math.ceil(filteredRows.length/rowsPerPage)||1;
    document.getElementById('pageInfo').textContent='Page '+currentPage+' of '+total;
    document.getElementById('prevBtn').disabled=currentPage===1;
    document.getElementById('nextBtn').disabled=currentPage>=total;
}
function sortTable(col,th){
    if(sortCol===col)sortAsc=!sortAsc;else{sortCol=col;sortAsc=true;}
    document.querySelectorAll('thead th').forEach(t=>t.className='');
    th.className=sortAsc?'asc':'desc';
    filteredRows.sort((a,b)=>{const va=a[col]||'',vb=b[col]||'',
        na=parseFloat(va),nb=parseFloat(vb);
        if(!isNaN(na)&&!isNaN(nb))return sortAsc?na-nb:nb-na;
        return sortAsc?va.localeCompare(vb):vb.localeCompare(va);});
    currentPage=1;renderTable();
}
function changePage(d){
    const total=Math.ceil(filteredRows.length/rowsPerPage)||1;
    currentPage=Math.min(Math.max(currentPage+d,1),total);renderTable();
}
document.getElementById('searchInput').oninput=function(){
    filteredRows=allRows.filter(r=>r.some(c=>c.toLowerCase().includes(this.value.toLowerCase())));
    currentPage=1;renderTable();
};
loadTable();setInterval(loadTable,5000);
</script>
</body>
</html>"""

def write_html_viewer():
    p = BASE_DIR / "view_csv.html"
    p.write_text(_HTML_VIEWER, encoding="utf-8")
    log.info(f"HTML dashboard written: {p}")

def export_csv_from_excel():
    try:
        if EXCEL_FILE.exists():
            df = pd.read_excel(EXCEL_FILE, engine="openpyxl")
            df.to_csv(BASE_DIR / "predicted_output.csv", index=False)
    except Exception as exc:
        SelfDebugger.capture(exc, "export_csv_from_excel")

def start_csv_exporter():
    def _loop():
        while True:
            time.sleep(4)
            export_csv_from_excel()
    threading.Thread(target=_loop, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def get_local_ip():
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "localhost"

def print_banner(profile, ml):
    print("\n" + "═" * 62)
    print("  ⚡  BATTERY SOC LIVE PREDICTION SYSTEM  ·  Multi-Model")
    print("═" * 62)
    print(f"  🔋  Battery : {profile['name']}")
    print(f"  🧠  Model   : {ml}")
    print("─" * 62)
    print("  [1]  AUTO    — Live Arduino sensor → ML prediction")
    print("  [2]  MANUAL  — Type SOC directly   (ESP32 relay test)")
    print("  [3]  CHANGE  — Switch battery type / model")
    print("  [4]  EXIT")
    print("═" * 62)


def main():
    log.info("Battery SOC Live System (Multi-Model) starting …")
    write_html_viewer()
    _ensure_excel()
    update_soc_json(0.0, "idle", "Not selected")

    safe_run(start_http_server, context="start_http_server")
    start_csv_exporter()

    ip = get_local_ip()
    print(f"\n  💡  Your PC IP address : {ip}")
    print(f"  📡  ESP32 SOC URL      : http://{ip}:{_http_port}/soc_feed.json")

    model_cache = preload_all_models()
    profile, model, feat_cols = select_battery(model_cache)

    while True:
        print_banner(profile, _model_label(profile))
        choice = input("  Enter choice [1/2/3/4]: ").strip()

        if choice == "1":
            if model is None:
                print("\n  ❌  Model not loaded — run training script first.\n")
                continue
            safe_run(auto_mode, model, feat_cols, profile, context="auto_mode")

        elif choice == "2":
            safe_run(manual_mode, profile, context="manual_mode")

        elif choice == "3":
            profile, model, feat_cols = select_battery(model_cache)

        elif choice == "4":
            print("\n  👋  Goodbye!\n")
            log.info("Program exited by user.")
            break

        else:
            print("  Invalid choice. Enter 1, 2, 3, or 4.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  Program interrupted. Goodbye!\n")
    except Exception as exc:
        SelfDebugger.capture(exc, "main()")
        sys.exit(1)
